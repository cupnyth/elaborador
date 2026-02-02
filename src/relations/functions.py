from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import os
from datetime import datetime
from openpyxl import load_workbook

class Lockerms:
    def __init__(self, nome_arquivo_modelo="modelo.xlsx"):
        # --- CONFIGURAÇÃO DE CAMINHOS ---
        self.pasta_modelos = "modelo_relacoes"
        self.modelo_excel_path = os.path.join(self.pasta_modelos, nome_arquivo_modelo)
        
        self.exportados_ms_path = "exportados_ms"
        self.exames_path = "exames"
        
        # Garante que as pastas existam
        os.makedirs(self.exportados_ms_path, exist_ok=True)
        os.makedirs(self.exames_path, exist_ok=True)

    def _atualizar_protocolo_msboi(self, colaborator: str, cpf: str):
        """
        Gerencia o Excel da MS BOI:
        - Se não existir arquivo hoje: Cria, põe a data (H5) e salva.
        - Se existir: Abre e adiciona o funcionario na proxima linha vazia (D e K).
        """
        try:
            # Define o nome do arquivo para HOJE
            agora = datetime.now()
            data_hoje_nome = agora.strftime("%d-%m-%Y")
            
            nome_arquivo_dia = f"Protocolo_MSBOI_{data_hoje_nome}.xlsx"
            caminho_final = os.path.join(self.exportados_ms_path, nome_arquivo_dia)

            wb = None
            ws = None

            # --- LÓGICA DE ABERTURA OU CRIAÇÃO ---
            if os.path.exists(caminho_final):
                # ARQUIVO JÁ EXISTE: Apenas carregamos
                print(f"[INFO] Atualizando protocolo existente: {nome_arquivo_dia}")
                wb = load_workbook(caminho_final)
                ws = wb.active
                # OBS: Não mexemos na data (H5) aqui, pois já foi posta na criação.
            else:
                # ARQUIVO NOVO: Carregamos do modelo
                print(f"[INFO] Criando novo protocolo para o dia: {nome_arquivo_dia}")
                if not os.path.exists(self.modelo_excel_path):
                    print(f"[ERRO] Modelo não encontrado em: {self.modelo_excel_path}")
                    return
                
                wb = load_workbook(self.modelo_excel_path)
                ws = wb.active
                
                # --- PREENCHE A DATA (SÓ UMA VEZ) ---
                # Data formatada na celula H5
                ws['H5'] = agora.strftime("%d/%m/%Y")

            # --- LÓGICA DE PREENCHIMENTO (LOOP) ---
            # Começa na linha 8 conforme solicitado
            linha_atual = 8
            
            # Vamos verificar se a coluna D (Nome) está vazia
            while True:
                celula_nome = ws[f'D{linha_atual}']
                
                # Se o valor for None ou vazio, achamos a linha livre
                if not celula_nome.value:
                    break
                
                # Se tem gente, desce uma linha
                linha_atual += 1

            # --- ESCREVENDO OS DADOS ---
            # Nome na Coluna D
            ws[f'D{linha_atual}'] = colaborator
            
            # CPF na Coluna K
            ws[f'K{linha_atual}'] = cpf
            
            # (Quantidade foi ignorada conforme pedido)

            # Salva
            wb.save(caminho_final)
            print(f"[SUCESSO] {colaborator} (CPF: {cpf}) salvo na linha {linha_atual}.")

        except Exception as e:
            print(f"[ERRO] Falha ao atualizar Excel MS BOI: {e}")

    def create_exam(self, enterprise: str, colaborator: str, exams: list, passw: str, cpf: str = ""):
        try:
            # --- 1. GERA O PDF (PADRÃO) ---
            os.makedirs(fr"{self.exames_path}\{enterprise}", exist_ok=True)
            os.makedirs(fr"temp", exist_ok=True)
            
            caminho_pdf = fr"{self.exames_path}\{enterprise}\{colaborator}.pdf"
            c = canvas.Canvas(caminho_pdf, pagesize=A4, encrypt=passw)
            
            for i, imagem in enumerate(exams):
                if os.path.exists(imagem):
                    c.drawImage(imagem, 0, 0, width=A4[0], height=A4[1], preserveAspectRatio=True)
                    if i < len(exams) - 1:
                        c.showPage()
            c.save()
            # ------------------------------

            # --- 2. VERIFICA SE É MS BOI ---
            empresa_alvo = "MS BOI COMERCIO E ABATE DE BOVINOS EIRELI"

            if enterprise.strip().upper() == empresa_alvo:
                # Dispara a lógica do Excel corrigida
                self._atualizar_protocolo_msboi(colaborator, cpf)
            
            return "Sucesso"

        except Exception as e:
            return f"Erro: {e}"